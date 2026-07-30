# Copyright 2025 SUNRIVERWOOD
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""
Rescore 分层得分分析模块 - 从已有的 rescore 评分结果文件中提取每道题目得分，
按 L1/L2/L3/L4/hard 分级统计，输出为 Excel 表格。

用法:
    # 分析指定目录（如 new-baseline）
    python -m evaluation.rescore_level_analysis --dir new-baseline

    # 分析所有目录
    python -m evaluation.rescore_level_analysis --all

    # 指定输出文件名
    python -m evaluation.rescore_level_analysis --dir old-baseline -o my_analysis.xlsx
"""

import argparse
import json
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
RESCORE_BASE_DIR = PROJECT_ROOT / "data" / "reports" / "rescore"
OUTPUT_DIR = PROJECT_ROOT / "data" / "reports" / "analysis"

# 难度分级定义
LEVEL_ORDER = ["L1", "L2", "L3", "L4", "hard", "Overall"]
MERGED_LEVEL_ORDER = ["L1", "L2", "L3", "L4", "Overall"]  # hard 按 difficulty 归入 L1~L4


def classify_level(record: dict) -> str:
    """
    根据 source_file 和 difficulty 字段确定题目的分级。
    - hard.json 中的题目统一归为 "hard"
    - 其他文件中的题目按 difficulty 字段分类 (L1/L2/L3/L4)
    """
    source_file = record.get("source_file", "")
    difficulty = record.get("difficulty", "L1")

    if source_file == "hard.json":
        return "hard"

    # L12.json 中包含 L1 和 L2，按实际 difficulty 区分
    if difficulty in ("L1", "L2", "L3", "L4"):
        return difficulty

    return difficulty


def load_score_file(filepath: Path) -> list[dict]:
    """加载单个评分结果文件"""
    records = []
    with open(filepath, 'r', encoding='utf-8') as f:
        for line in f:
            if line.strip():
                record = json.loads(line)
                records.append(record)
    return records


def discover_score_files(base_dir: Path) -> dict[str, list[Path]]:
    """发现目录下（含子目录）的所有评分文件，按方法名分组"""
    method_files = defaultdict(list)
    for f in sorted(base_dir.rglob("*_scores_*.jsonl")):
        # 从文件名提取方法名: {method}_scores_{timestamp}.jsonl
        parts = f.stem.split("_scores_")
        if len(parts) == 2:
            method_name = parts[0]
            method_files[method_name].append(f)
    return dict(method_files)


def find_analysis_dirs(base_dir: Path) -> list[Path]:
    """
    递归发现所有包含评分文件的目录。
    对于含子目录的情况（如 unstrict/ablation_*），返回叶子级目录。
    对于直接含评分文件的目录（如 new-baseline），返回该目录本身。
    """
    result = []
    # 先检查当前目录是否直接含评分文件
    direct_files = list(base_dir.glob("*_scores_*.jsonl"))
    if direct_files:
        result.append(base_dir)
    # 递归检查子目录
    for d in sorted(base_dir.iterdir()):
        if d.is_dir():
            result.extend(find_analysis_dirs(d))
    return result


def analyze_directory(dir_path: Path) -> dict:
    """
    分析一个目录，返回结构化的分析结果。

    返回:
        {
            "dir_name": str,
            "methods": {
                method_name: {
                    "records": [record, ...],
                    "level_stats": {
                        "L1": {"avg": float, "count": int, "min": float, "max": float, "scores": [...]},
                        ...
                    },
                    "overall": {"avg": float, ...}
                }
            }
        }
    """
    # 只搜索当前目录（不递归），递归由 find_analysis_dirs 控制
    method_files = defaultdict(list)
    for f in sorted(dir_path.glob("*_scores_*.jsonl")):
        parts = f.stem.split("_scores_")
        if len(parts) == 2:
            method_files[parts[0]].append(f)
    if not method_files:
        return None

    result = {"dir_name": dir_path.name, "methods": {}}

    for method, files in sorted(method_files.items()):
        # 取最新的文件
        latest_file = sorted(files)[-1]
        records = load_score_file(latest_file)

        # 按 level 分组统计
        level_scores = defaultdict(list)
        all_valid_records = []

        for r in records:
            is_error = r.get("scores", {}).get("error", False)
            score = r.get("overall_score", 0.0)
            level = classify_level(r)

            enriched = {
                "question_id": r.get("question_id"),
                "question": r.get("question", "")[:80],
                "level": level,
                "difficulty": r.get("difficulty", ""),
                "source_file": r.get("source_file", ""),
                "type": r.get("type", ""),
                "domain": r.get("domain", ""),
                "overall_score": score,
                "is_error": is_error,
                "method": method,
            }
            all_valid_records.append(enriched)

            if not is_error:
                level_scores[level].append(score)

        # 计算各 level 统计
        level_stats = {}
        all_scores = []
        for level in LEVEL_ORDER[:-1]:  # 不含 Overall
            scores = level_scores.get(level, [])
            if scores:
                level_stats[level] = {
                    "avg": sum(scores) / len(scores),
                    "count": len(scores),
                    "min": min(scores),
                    "max": max(scores),
                    "scores": scores,
                }
                all_scores.extend(scores)
            else:
                level_stats[level] = {"avg": 0, "count": 0, "min": 0, "max": 0, "scores": []}

        # Overall
        if all_scores:
            level_stats["Overall"] = {
                "avg": sum(all_scores) / len(all_scores),
                "count": len(all_scores),
                "min": min(all_scores),
                "max": max(all_scores),
            }
        else:
            level_stats["Overall"] = {"avg": 0, "count": 0, "min": 0, "max": 0}

        # ── 合并视图: hard 按 difficulty 归入 L1~L4 ──
        merged_scores = defaultdict(list)
        for r in all_valid_records:
            if r["is_error"]:
                continue
            # hard 题按 difficulty 字段归入 L1~L4
            merged_level = r["difficulty"] if r["difficulty"] in ("L1", "L2", "L3", "L4") else "L1"
            merged_scores[merged_level].append(r["overall_score"])

        merged_stats = {}
        merged_all = []
        for level in MERGED_LEVEL_ORDER[:-1]:
            scores = merged_scores.get(level, [])
            if scores:
                merged_stats[level] = {
                    "avg": sum(scores) / len(scores),
                    "count": len(scores),
                    "min": min(scores),
                    "max": max(scores),
                }
                merged_all.extend(scores)
            else:
                merged_stats[level] = {"avg": 0, "count": 0, "min": 0, "max": 0}

        if merged_all:
            merged_stats["Overall"] = {
                "avg": sum(merged_all) / len(merged_all),
                "count": len(merged_all),
                "min": min(merged_all),
                "max": max(merged_all),
            }
        else:
            merged_stats["Overall"] = {"avg": 0, "count": 0, "min": 0, "max": 0}

        result["methods"][method] = {
            "records": all_valid_records,
            "level_stats": level_stats,
            "merged_level_stats": merged_stats,
        }

    return result


# ─── Excel 样式定义 ───────────────────────────────────────────────
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
LEVEL_FILLS = {
    "L1": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "L2": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "L3": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "L4": PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid"),
    "hard": PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
    "Overall": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
NUMBER_FMT = "0.0000"
PCT_FMT = "0.00%"


def _style_header_row(ws, row, max_col):
    """为表头行应用样式"""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def write_summary_sheet(wb: Workbook, analysis: dict):
    """
    写入汇总 Sheet：方法 × 难度等级 的平均分矩阵
    """
    ws = wb.active
    ws.title = "Summary"

    methods = sorted(analysis["methods"].keys())

    # ── 表头 ──
    headers = ["Method"] + [f"{l} Avg" for l in LEVEL_ORDER] + [f"{l} Count" for l in LEVEL_ORDER[:-1]]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    # ── 数据行 ──
    for row_idx, method in enumerate(methods, 2):
        stats = analysis["methods"][method]["level_stats"]
        ws.cell(row=row_idx, column=1, value=method).border = THIN_BORDER

        # 平均分列
        for col_offset, level in enumerate(LEVEL_ORDER):
            cell = ws.cell(row=row_idx, column=2 + col_offset, value=stats[level]["avg"])
            cell.number_format = NUMBER_FMT
            cell.border = THIN_BORDER
            if level in LEVEL_FILLS:
                cell.fill = LEVEL_FILLS[level]
            cell.alignment = Alignment(horizontal="center")

        # 数量列
        base_col = 2 + len(LEVEL_ORDER)
        for col_offset, level in enumerate(LEVEL_ORDER[:-1]):
            cell = ws.cell(row=row_idx, column=base_col + col_offset, value=stats[level]["count"])
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

    # 列宽
    ws.column_dimensions["A"].width = 35
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # 冻结首行
    ws.freeze_panes = "B2"


def write_detail_sheet(wb: Workbook, analysis: dict, method: str):
    """为每个方法写入详细得分 Sheet"""
    data = analysis["methods"][method]
    records = sorted(data["records"], key=lambda r: r["question_id"])

    ws = wb.create_sheet(title=method[:31])  # Sheet 名最长 31 字符

    headers = ["QID", "Level", "Difficulty", "Source", "Type", "Domain", "Score", "Error", "Question"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    for row_idx, r in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=r["question_id"]).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=r["level"]).border = THIN_BORDER
        ws.cell(row=row_idx, column=3, value=r["difficulty"]).border = THIN_BORDER
        ws.cell(row=row_idx, column=4, value=r["source_file"]).border = THIN_BORDER
        ws.cell(row=row_idx, column=5, value=r["type"]).border = THIN_BORDER
        ws.cell(row=row_idx, column=6, value=r["domain"]).border = THIN_BORDER

        score_cell = ws.cell(row=row_idx, column=7, value=r["overall_score"])
        score_cell.number_format = NUMBER_FMT
        score_cell.border = THIN_BORDER
        if r["level"] in LEVEL_FILLS:
            score_cell.fill = LEVEL_FILLS[r["level"]]

        ws.cell(row=row_idx, column=8, value="Yes" if r["is_error"] else "").border = THIN_BORDER
        ws.cell(row=row_idx, column=9, value=r["question"]).border = THIN_BORDER

    # 列宽
    col_widths = [6, 8, 10, 12, 16, 14, 10, 6, 60]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"


def write_level_comparison_sheet(wb: Workbook, analysis: dict):
    """写入各 level 对比表：行=方法，列=level，值=avg score，带颜色条"""
    ws = wb.create_sheet(title="Level Comparison")

    methods = sorted(analysis["methods"].keys())

    # 表头：Method, L1_avg, L1_min, L1_max, L2_avg, ...
    headers = ["Method"]
    for level in LEVEL_ORDER[:-1]:
        headers.extend([f"{level}_avg", f"{level}_min", f"{level}_max"])
    headers.extend(["Overall_avg", "Overall_min", "Overall_max"])
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    for row_idx, method in enumerate(methods, 2):
        stats = analysis["methods"][method]["level_stats"]
        ws.cell(row=row_idx, column=1, value=method).border = THIN_BORDER
        col = 2
        for level in LEVEL_ORDER:
            s = stats[level]
            for val in [s["avg"], s.get("min", 0), s.get("max", 0)]:
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.number_format = NUMBER_FMT
                cell.border = THIN_BORDER
                if level in LEVEL_FILLS:
                    cell.fill = LEVEL_FILLS[level]
                cell.alignment = Alignment(horizontal="center")
                col += 1

    ws.column_dimensions["A"].width = 35
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12

    ws.freeze_panes = "B2"


def write_merged_summary_sheet(wb: Workbook, analysis: dict):
    """
    写入合并视图 Sheet：hard 题按 difficulty 归入 L1~L4
    """
    ws = wb.create_sheet(title="Summary (Merged)")

    methods = sorted(analysis["methods"].keys())

    # 表头
    headers = ["Method"] + [f"{l} Avg" for l in MERGED_LEVEL_ORDER] + [f"{l} Count" for l in MERGED_LEVEL_ORDER[:-1]]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    for row_idx, method in enumerate(methods, 2):
        stats = analysis["methods"][method]["merged_level_stats"]
        ws.cell(row=row_idx, column=1, value=method).border = THIN_BORDER

        for col_offset, level in enumerate(MERGED_LEVEL_ORDER):
            cell = ws.cell(row=row_idx, column=2 + col_offset, value=stats[level]["avg"])
            cell.number_format = NUMBER_FMT
            cell.border = THIN_BORDER
            if level in LEVEL_FILLS:
                cell.fill = LEVEL_FILLS[level]
            cell.alignment = Alignment(horizontal="center")

        base_col = 2 + len(MERGED_LEVEL_ORDER)
        for col_offset, level in enumerate(MERGED_LEVEL_ORDER[:-1]):
            cell = ws.cell(row=row_idx, column=base_col + col_offset, value=stats[level]["count"])
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 35
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.freeze_panes = "B2"


def export_to_excel(analysis: dict, output_path: Path):
    """将分析结果导出为 Excel 文件"""
    wb = Workbook()

    # Sheet 1: Summary (原始视图: L1~L4 + hard)
    write_summary_sheet(wb, analysis)

    # Sheet 2: Summary Merged (合并视图: hard 归入 L1~L4)
    write_merged_summary_sheet(wb, analysis)

    # Sheet 3: Level Comparison (详细 min/max)
    write_level_comparison_sheet(wb, analysis)

    # Sheet 4+: 每个方法的逐题详情
    for method in sorted(analysis["methods"].keys()):
        write_detail_sheet(wb, analysis, method)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"\n✅ Excel 已保存: {output_path}")


def _print_table(methods, stats_key, level_order, analysis, title):
    """通用终端表格打印"""
    print(f"\n{'='*100}")
    print(f"  {title}")
    print(f"{'='*100}")

    header = f"{'Method':<35}"
    for level in level_order:
        header += f" {level:>8}"
    print(header)
    print(f"{'-'*100}")

    for method in methods:
        stats = analysis["methods"][method][stats_key]
        row = f"{method:<35}"
        for level in level_order:
            avg = stats[level]["avg"]
            count = stats[level]["count"]
            row += f" {avg:>7.4f}" if count > 0 else f" {'N/A':>7}"
        print(row)

    print(f"{'-'*100}")
    row = f"{'Count':<35}"
    first_method = methods[0] if methods else None
    if first_method:
        stats = analysis["methods"][first_method][stats_key]
        for level in level_order:
            row += f" {stats[level]['count']:>8}"
        print(row)
    print(f"{'='*100}")


def print_summary_table(analysis: dict):
    """在终端打印汇总表（两种视图）"""
    methods = sorted(analysis["methods"].keys())
    dir_name = analysis["dir_name"]

    # 视图1: 原始分类 (L1~L4 + hard)
    _print_table(methods, "level_stats", LEVEL_ORDER, analysis,
                 f"Rescore 分层得分 [{dir_name}] — 原始分类")

    # 视图2: 合并分类 (hard 归入 L1~L4)
    _print_table(methods, "merged_level_stats", MERGED_LEVEL_ORDER, analysis,
                 f"Rescore 分层得分 [{dir_name}] — 合并视图 (hard→L1~L4)")


def main():
    parser = argparse.ArgumentParser(description="Rescore 分层得分分析")
    parser.add_argument("--dir", type=str, default=None,
                        help="指定要分析的子目录名称，如: new-baseline, old-baseline")
    parser.add_argument("--all", action="store_true",
                        help="分析所有子目录")
    parser.add_argument("-o", "--output", type=str, default=None,
                        help="指定输出 Excel 文件名")
    args = parser.parse_args()

    if not args.dir and not args.all:
        # 默认列出可用目录（递归）
        print("可用的 rescore 结果目录:")
        all_dirs = find_analysis_dirs(RESCORE_BASE_DIR)
        for d in all_dirs:
            rel = d.relative_to(RESCORE_BASE_DIR)
            n_files = len(list(d.glob("*_scores_*.jsonl")))
            print(f"  - {rel} ({n_files} 个评分文件)")
        print(f"\n共 {len(all_dirs)} 个目录。使用 --dir <name> 指定目录，或 --all 分析全部。")
        print("提示: 支持嵌套路径，如 --dir unstrict\\ablation_reasoning_gnn_direct")
        return

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if args.all:
        # 递归分析所有含评分文件的目录
        all_dirs = find_analysis_dirs(RESCORE_BASE_DIR)
        print(f"发现 {len(all_dirs)} 个待分析目录")
        for d in all_dirs:
            analysis = analyze_directory(d)
            if analysis:
                rel_name = str(d.relative_to(RESCORE_BASE_DIR)).replace("\\", "_").replace("/", "_")
                analysis["dir_name"] = str(d.relative_to(RESCORE_BASE_DIR))
                print_summary_table(analysis)
                out_name = f"rescore_level_{rel_name}_{timestamp}.xlsx"
                export_to_excel(analysis, OUTPUT_DIR / out_name)
    else:
        dir_path = RESCORE_BASE_DIR / args.dir
        if not dir_path.exists():
            print(f"❌ 目录不存在: {dir_path}")
            sys.exit(1)

        # 检查是否含子目录
        sub_dirs = find_analysis_dirs(dir_path)
        if not sub_dirs:
            print(f"❌ 目录 {args.dir} 中未找到评分文件")
            sys.exit(1)

        for d in sub_dirs:
            analysis = analyze_directory(d)
            if analysis:
                rel_name = str(d.relative_to(RESCORE_BASE_DIR)).replace("\\", "_").replace("/", "_")
                analysis["dir_name"] = str(d.relative_to(RESCORE_BASE_DIR))
                print_summary_table(analysis)
                out_name = args.output or f"rescore_level_{rel_name}_{timestamp}.xlsx"
                export_to_excel(analysis, OUTPUT_DIR / out_name)


if __name__ == "__main__":
    main()
